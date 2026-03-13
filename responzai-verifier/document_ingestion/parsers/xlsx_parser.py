# document_ingestion/parsers/xlsx_parser.py

from openpyxl import load_workbook

async def parse_xlsx(file_path: str) -> dict:
    """
    Extrahiert Text aus Excel-Tabellen.

    Warum Excel prüfen?
    GUIDES-Vorlagen enthalten manchmal Checklisten oder
    Bewertungsmatrizen als Excel-Tabellen. Diese enthalten
    Behauptungen, die geprüft werden müssen.

    Jedes Blatt (Sheet) wird einzeln verarbeitet.
    Die Zellinhalte werden als Tabelle formatiert.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sections = []
    full_text = ""

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_text = f"Blatt: {sheet_name}\n"

        for row in ws.iter_rows(values_only=True):
            # Leere Zellen überspringen
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(cells):
                sheet_text += " | ".join(cells) + "\n"

        sections.append({
            "heading": f"Blatt: {sheet_name}",
            "text": sheet_text
        })
        full_text += sheet_text + "\n\n"

    wb.close()

    return {
        "text": full_text,
        "sections": sections,
        "pages": len(wb.sheetnames)
    }
